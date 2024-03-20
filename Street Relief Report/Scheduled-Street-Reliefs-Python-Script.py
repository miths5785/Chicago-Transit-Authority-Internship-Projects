# import necessary modules

import os
import openpyxl
import pandas as pd
import seaborn as sns
from win32com import client
from datetime import datetime
from urllib.parse import urlparse
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font, Alignment, Border, Side

# reading the csv file obtained from the OIG export (Enter the complete path of the file)

# Ex. C:\Users\MBhutada.int\OneDrive - Chicago Transit Authority\Documents\Python Scripts\Street_Relief_P-11.csv

path = input("Please enter the complete path of the csv file: ")

df = pd.read_csv(path)

# extracting the file name

a = urlparse(path)
filename = os.path.basename(a.path)

# obtaining the pick name

pick = df['Booking Name'][0] + " Pick"
pick = pick.split(' - ')[1]

# removing leading and trailing whitespaces from the column names in the dataframe

df.columns = df.columns.str.strip()

# condtional filtering to include only the street reliefs

df = df[(df['Relief Scheduled'] == 1) & (df['Street Relief'] == 1) & (df['Relief at Block Start'] == 0) & (df['Relief at Block End'] == 0)]

# replacing special characters from the 'Relief Time' column with appropriate characters

df['Relief Time'] = df['Relief Time'].str.replace(' ', '')
df['Relief Time'] = df['Relief Time'].str.replace('+', '')
df['Relief Time'] = df['Relief Time'].str.replace("'", ':')

# resetting the index of the dataframe and dropping the index column created

df = df.reset_index()
df.drop('index', axis = 1, inplace = True)

# dropping the duplicate rows from the dataframe and keeping the first copy only

df = df.drop_duplicates(keep = 'first')

# resetting the index of the dataframe and dropping the index column created

df = df.reset_index()
df.drop('index', axis = 1, inplace = True)

# converting the time in the 'Relief Time' column to the 24-hour format

def date_convert(date_to_convert):
     return datetime.strptime(date_to_convert, "%H:%M").strftime("%H:%M")

df['Relief Time'] = df['Relief Time'].apply(date_convert)

# retrieving the Hour and Minute portion from the 'Relief Time' column

hour = df['Relief Time'].str.split(":").str[0]
minute = df['Relief Time'].str.split(":").str[1]

# applying condtional formatting to the Hour and Minute portion to get the desired time-format in the report

for i in range(len(minute)):
    if int(minute[i]) >= 1 and int(minute[i]) <= 30:
        minute[i] = 0.00
    
    elif int(minute[i]) >= 31 and int(minute[i]) <= 59:
        minute[i] = 0.30
    
    elif int(minute[i]) == 00:
        minute[i] = 0.30
        hour[i] = int(hour[i]) - 1
    
    else:
        pass

# converting Hour and Minute into suitable datatypes

minute = minute.astype(float)
hour = hour.astype(int)

# obtaining the desired time values in the format '> HH:00' and '> HH:30' respectively for the report

time = hour + minute

for e in range(len(time)):
    time[e] = str(time[e]).replace(".", ":") + "0"
    
    if len(time[e]) == 4:
        time[e] = "0" + time[e]
    
    time[e] = "> " + time[e]

# creating a new column 'Half-Hour' column in the dataframe and storing time values in it

df['Half-Hour'] = time

# creating a new column 'Weight' in the dataframe and assigning weights to
# it based on the number of operating days ('Relief Operating Days' column)

def define_weight(day_to_convert):
    if (len(day_to_convert) == 1 or len(day_to_convert) == 5) and day_to_convert in ['s', 'a', 'muwtf']:
        return 1
    
    else:
        return len(day_to_convert) * 0.2

df['Weight'] = df['Relief Operating Days'].apply(define_weight)

# removing leading and trailing whitespaces from "object" type column values

obj_columns = list(df.select_dtypes(include = ['object']))
for i in obj_columns:
    df[i] = df[i].astype(str).str.strip()

# converting 'Weight' column to numerical type

df['Weight'] = pd.to_numeric(df['Weight'])

# dividing the dataframe into 3 day-types dataframes as required by the report

df_weekday = df[df['Crew Schedule Type'] == 'Weekday']
df_saturday = df[df['Crew Schedule Type'] == 'Saturday']
df_sunday = df[df['Crew Schedule Type'] == 'Sunday']

# assigning variables

len_df_list = 3
df_to_pivot_table_list = []

df_all_days = [df_weekday, df_saturday, df_sunday]

# converting the 3 day-type dataframes into respective pivot tables

for i in range(len_df_list):
    x = df_all_days[i]
    x = x.pivot_table(index = ['Half-Hour'],
                      columns = ['Block Garage'],
                      values = 'Weight',
                      aggfunc = 'sum',
                      fill_value = 0,
                      margins = True,
                      margins_name = 'Grand Total')

    order = ['F','P','5','K','6','7','1', 'Grand Total']

    x = x.reindex(columns = order)
    
    # appending the obtained pivot tables to df_to_pivot_table_list 
    
    df_to_pivot_table_list.append(x)

# defining a color map (red) for the pivot table cells 

cm = sns.light_palette("#FF3333", as_cmap = True)

# defining an excel file based in the path variable

path = path[:-4] + ".xlsx" 

# defining a writer class to write pivot tables to an excel file

writer = pd.ExcelWriter(path, engine='xlsxwriter')   
workbook = writer.book
worksheet = workbook.add_worksheet('Report')

# initializing variables

startrow = 10
startcol = 1
total_cols = [] # list for storing the grand-total column in pivot table of each day-type
total_rows = [] # list for storing the grand-total row in pivot table of each day-type
total_final = [] # list for storing the grand-total number in pivot table of each day-type

# writing the pivot tables obtained into the excel sheet

for i, j in zip(df_to_pivot_table_list, range(len_df_list)):
    final_fig = [] # list for storing the grand-total number of the current pivot table in consideration
    
    half_hour_list = {'> 05:30' : 0, '> 06:00' : 0, '> 06:30' : 0, '> 07:00' : 0, '> 07:30' : 0, '> 08:00' : 0,
                      '> 08:30' : 0, '> 09:00' : 0, '> 09:30' : 0, '> 10:00' : 0, '> 10:30' : 0, '> 11:00' : 0,
                      '> 11:30' : 0, '> 12:00' : 0, '> 12:30' : 0, '> 13:00' : 0, '> 13:30' : 0, '> 14:00' : 0,
                      '> 14:30' : 0, '> 15:00' : 0, '> 15:30' : 0, '> 16:00' : 0, '> 16:30' : 0, '> 17:00' : 0,
                      '> 17:30' : 0, '> 18:00' : 0, '> 18:30' : 0, '> 19:00' : 0, '> 19:30' : 0, '> 20:00' : 0,
                      '> 20:30' : 0, '> 21:00' : 0, '> 21:30' : 0, '> 22:00' : 0}
    
    # retrieving the respective grand-total attributes for the pivot table in consideration
    
    grand_total_column = i.iloc[:-1,-1]
    grand_total_row = i.iloc[-1,:-1]
    grand_total_final = i.iloc[-1,-1]
    final_fig.append(grand_total_final)
    
    # retrieving the pivot table without the grand-total attributes
    
    pivot_without_totals = i.iloc[:-1,:-1]
    
    # obtaining the counts of Half-Hour values
    
    for h in pivot_without_totals.index:
        if h in half_hour_list:
            half_hour_list[h] = half_hour_list[h] + 1
        else:
            pass

    #  defining the Half-Hour index for the pivot table if not already present
    
    for h in half_hour_list:
        if half_hour_list[h] == 0:
            pivot_without_totals.loc[h] = [0, 0, 0, 0, 0, 0, 0]
            grand_total_column[h] = 0
            
    # appending the grand-total attributes of the current pivot table to the defined lists
            
    total_cols.append(grand_total_column)
    total_rows.append(grand_total_row)
    total_final.append(grand_total_final)
    
    # converting the grand-total attributes of the current pivot table to dataframes
    
    grand_total_column = pd.DataFrame(grand_total_column)
    grand_total_row = pd.DataFrame(grand_total_row)
    grand_total_row.rename(columns = {'Grand Total' : ''}, inplace = True)
    grand_total_final = pd.DataFrame(final_fig, columns = ["Grand Total"])
    
    # converting the respective pivot table and grand-total column to styler objects by providing the color gradient using the color map defined above
    
    pivot_without_totals = pivot_without_totals.style.background_gradient(cmap = cm, axis = None)
    grand_total_column = grand_total_column.style.background_gradient(cmap = cm, axis = 0)
    
    df_to_pivot_table_list[j] = pivot_without_totals
    
    # writing the respective objects to the excel sheet
    
    pivot_without_totals.to_excel(writer, sheet_name = 'Report', startrow = startrow , startcol = startcol, header = ['Forest Park','North Park','Chicago','Kedzie','74th','77th','103rd'])
    grand_total_column.to_excel(writer, sheet_name = 'Report', startrow = startrow , startcol = startcol + 9, index = False, header = ['Total'])
    grand_total_row.T.to_excel(writer, sheet_name = 'Report', startrow = startrow + 35 , startcol = startcol, header = False)
    grand_total_final.to_excel(writer, sheet_name = 'Report', startrow = startrow + 35 , startcol = startcol + 9, index = False, header = False)
    
    startrow = startrow
    startcol = startcol + 12

# defining list for storing the percentage pivot tables

pivot_table_by_percentage = []

# computing the percentage pivot table

for i, n in zip(df_to_pivot_table_list, total_rows):
    for j in i.data.index:
        i.data.loc[j] = round(i.data.loc[j]*100/n, 1)
    
    # appending the percentage pivot table to list defined above
    
    pivot_table_by_percentage.append(i.data)

# initializing variables

startrow = 52
startcol = 1

for i, j in zip(pivot_table_by_percentage, range(len_df_list)):
    
    # converting the respective percentage pivot table to styler objects by providing the color gradient using the color map defined above
    
    i = i.style.background_gradient(cmap = cm, axis = None)
    
    # calculating the percentage in the grand-total column of the respective pivot table
    
    grand_total_column_percent_out_of_total = round(total_cols[j] * 100/total_final[j], 2)
    
    # calculating the cumulative percentage sum from the grand-total column of the respective pivot table
    
    grand_total_column_percent_out_of_total_cum_sum = grand_total_column_percent_out_of_total.cumsum()
    
    # rounding the obtained columns to 1 decimal place
    
    grand_total_column_percent_out_of_total = round(grand_total_column_percent_out_of_total, 1)
    grand_total_column_percent_out_of_total_cum_sum = round(grand_total_column_percent_out_of_total_cum_sum, 1)
    
    # converting the grand-total column by percentage to a styler object
    
    grand_total_column_percent_out_of_total = pd.DataFrame(grand_total_column_percent_out_of_total).style.background_gradient(cmap = cm, axis = None)
    
    # writing the respective objects to the excel sheet
    
    i.to_excel(writer, sheet_name = 'Report',startrow = startrow , startcol = startcol, header = ['Forest Park','North Park','Chicago','Kedzie','74th','77th','103rd'])
    grand_total_column_percent_out_of_total.to_excel(writer, sheet_name = 'Report',startrow = startrow , startcol = startcol + 9, index = False, header = ['Total'])
    grand_total_column_percent_out_of_total_cum_sum.to_excel(writer, sheet_name = 'Report',startrow = startrow , startcol = startcol + 10, index = False, header = ["Cumul."])
    
    startrow = startrow
    startcol = startcol + 12

# closing the writer object
    
writer.close() 

# open excel workbook

workbook = openpyxl.load_workbook(path)

# select the worksheet

worksheet = workbook.active

# setting the font and alignment of each cell in the worksheet

for row in range(1, worksheet.max_row + 1):
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row, col)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell.font = Font(name = 'Calibri', size = 11, bold = False, color = '000000')

# defining font and alignment for pivot table names in the excel file and applying it to the respective cells   
        
font = Font(name = 'Calibri', size = 18, bold = True)
alignment = Alignment(horizontal = 'center',
                      vertical = 'center')

worksheet['F9'] = "Weekday"
worksheet['F9'].font = font
worksheet['F9'].alignment = alignment

worksheet['R9'] = "Saturday"
worksheet['R9'].font = font
worksheet['R9'].alignment = alignment

worksheet['AD9'] = "Sunday"
worksheet['AD9'].font = font
worksheet['AD9'].alignment = alignment

# defining the font for the excel file header and applying it to the respective cell

font = Font(name = 'Calibri', size = 36, bold = True)

worksheet['R4'] = "SCHEDULED STREET RELIEFS"
worksheet['R4'].font = font
worksheet['R4'].alignment = alignment

# defining the font for the excel file pick and applying it to the respective cell

font = Font(name = 'Calibri', size = 20, bold = False)

worksheet['R6'] = pick
worksheet['R6'].font = font
worksheet['R6'].alignment = alignment

# defining the font for the percentage pivot tables header and applying it to the respective cell

font = Font(name = 'Calibri', size = 16, bold = True, italic = True)

worksheet['R50'] = 'As Percentage by Garage'
worksheet['R50'].font = font
worksheet['R50'].alignment = alignment

# defining the border style and the font

thin = Side(border_style = "thin")
font = Font(name = 'Calibri', size = 11, bold = True)

# removing borders from the Half-Hour Column and setting its font
        
column = 2

for i in range(3):    
    for col in worksheet.iter_cols(min_row = 12, min_col = column, max_col = column, max_row = 87):
        for cell in col:
            cell.border = Border(top = None, left = None, right = None, bottom = None)
            cell.font = font
        column = column + 12

# adding top border to the "Total's" Row in the top 3 tables and setting its font

min_col = 2
max_col = 11

for i in range(3):
    for row in worksheet.iter_rows(min_row = 46, min_col = min_col, max_col = max_col, max_row = 46):
        for cell in row:
            cell.border = Border(top = thin, left = None, right = None, bottom = None)
            cell.font = font
        min_col = min_col + 12
        max_col = max_col + 12

# adding bottom border to the column headers of each tables, setting the alignment and font as well

# defining the alignment, font, and border style

alignment = Alignment(horizontal = 'center',
                      vertical = 'center', 
                      wrap_text = True)
font = Font(name = 'Calibri', size = 11, bold = False)
thin = Side(border_style = "thin")

min_col = 2
max_col = 11
counter = 0

for i in range(6):
    if counter < 3:
        for row in worksheet.iter_rows(min_row = 11, min_col = min_col, max_col = max_col, max_row = 11):
            for cell in row:
                cell.alignment = alignment
                cell.font = font
                cell.border = Border(top = None, left = None, right = None, bottom = thin)
            min_col = min_col + 12
            max_col = max_col + 12
            counter = counter + 1
        
    else:
        if counter == 3:
            min_col = min_col - 12
            max_col = max_col - 11
        else:
            pass
        for row in worksheet.iter_rows(min_row = 53, min_col = min_col, max_col = max_col, max_row = 53):
            for cell in row:
                cell.alignment = alignment
                cell.font = font
                cell.border = Border(top = None, left = None, right = None, bottom = thin)
            min_col = min_col - 12
            max_col = max_col - 12
            counter = counter + 1
            
# applying italics font styling to the 'Cumul.' column values font, and aligning it

font = Font(italic = True)
alignment = Alignment(horizontal = 'right',
                      vertical = 'center')

column = 12

for i in range(3):    
    for col in worksheet.iter_cols(min_row = 53, min_col = column, max_col = column, max_row = 87):
        for cell in col:
            cell.font = font
            cell.alignment = alignment
        column = column + 12
            
# making the text bold in the "Total's" Column of each table

font = Font(name = 'Calibri', size = 11, bold = True, italic = False)
column = 11

for i in range(3):    
    for col in worksheet.iter_cols(min_row = 11, min_col = column, max_col = column, max_row = 87):
        for cell in col:
            cell.font = font
        column = column + 12
        
#  making the "Cumul." column header bold
        
cumul_cell_list = ['L53', 'X53', 'AJ53']

for cell in cumul_cell_list:
    worksheet[cell].font = font
    
# defining the number format for percentage pivot tables

for row in range(54, worksheet.max_row+1):
    for col in range(3, worksheet.max_column+1):
        cell = worksheet.cell(row, col)
        cell.number_format = '0.0'
        
# replacing all non-bold zero values in the pivot tables with ''

for row in range(1, worksheet.max_row+1):
    for col in range(1, worksheet.max_column+1):
        cell = worksheet.cell(row, col)
        if cell.value in [0, 0.0, '0', '0.0'] and cell.font.bold != True:
            cell.value = ''
        else:
            pass

# deleting unnecessary rows in the excel worksheet

rows  = [1, 1, 1, 2, 3, 3, 4, 40, 40, 42, 42]

for row in rows:
    worksheet.delete_rows(idx = row)

# setting up row heights

rows = [1, 2, 3, 4, 40, 41, 42]
row_heights = [46.2, 45, 23.4, 27.6, 26.3, 24.8, 27.6]

for row, row_height in zip(rows, row_heights):
    worksheet.row_dimensions[row].height = row_height

# setting up the cell dimensions in the entire worksheet

lower_limit = 5
upper_limit = 40

for i in range(2):
    for row in range(lower_limit, upper_limit):
        worksheet.row_dimensions[row].height = 14.4

    lower_limit = upper_limit + 4 
    upper_limit = 77
    
columns = ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AI', 'AJ']

for column in columns:
    worksheet.column_dimensions[column].width = 7

columns = ['B', 'N', 'Z']

for column in columns:
    worksheet.column_dimensions[column].width = 8.67
    
columns = ['J', 'V', 'AH']

for column in columns:
    worksheet.column_dimensions[column].width = 3.56

# defining font for column headers for all pivot tables    

font = Font(name = 'Calibri', size = 10)  
min_col = 2
max_col = 35
row = 4

for i in range(2):
    for row in worksheet.iter_rows(min_row = row, min_col = min_col, max_col = max_col, max_row = row):
        for cell in row:
            cell.font = font
    row = 42

# defining font for "Total's" column header for all pivot tables
    
font = Font(name = 'Calibri', size = 10, bold = True)      
Totals_cell_list = ['K4', 'K42', 'W4', 'W42', 'AI4', 'AI42']

for cell in Totals_cell_list:
    worksheet[cell].font = font    
    
# save changes to the workbook

workbook.save(path)

# open Microsoft Excel

excel = client.Dispatch("Excel.Application")
excel.Visible = False

# read excel file

sheets = excel.Workbooks.Open(path)
work_sheets = sheets.ActiveSheet

# set the desired print settings

work_sheets.PageSetup.Orientation = 2
work_sheets.PageSetup.Zoom = False
work_sheets.PageSetup.FitToPagesTall = False
work_sheets.PageSetup.FitToPagesWide = 1
work_sheets.PageSetup.TopMargin = 23
work_sheets.PageSetup.BottomMargin = 10
work_sheets.PageSetup.RightMargin = 65
work_sheets.PageSetup.LeftMargin = 65

# convert into PDF File

path = path[:-4] + "pdf" 
work_sheets.ExportAsFixedFormat(0, path)
sheets.Close(SaveChanges = False)
excel.Quit()

# success/failure message

if os.path.isfile(path) == True:
    print("Street Relief Report created successfully!!!")

else:
    print("Something went wrong..")

input("Press any key to quit")